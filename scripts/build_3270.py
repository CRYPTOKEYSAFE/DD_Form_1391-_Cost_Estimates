import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

B = Font(bold=True)
THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR = PatternFill("solid", fgColor="D9D9D9")
LOCK = PatternFill("solid", fgColor="FFF2CC")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# COVER
c = wb.create_sheet("COVER")
c["A1"] = "DD FORM 1391 COST ESTIMATE"; c["A1"].font = Font(bold=True, size=14)
rows = [
    ("Building", "SCH-3270"),
    ("iNFADS Description", "AUTO ORGANIZATIONAL SHOP CAB"),
    ("Primary CCN", 21451),
    ("RPUID", 1174058),
    ("PAX ID", 387568),
    ("Fi Web Project Number", "BU26PPE74M"),
    ("Installation/UIC", "M67400 (SH)"),
    ("Location", "MARINE CORPS BASE, CAMP SMEDLEY BUTLER (CAMP SCHWAB-6009), HENOKO OKINAWA, JAPAN"),
    ("Project Title", "Repair Bldg. 3270"),
    ("Fund Type", "O&MMC"),
    ("Acquisition Method", "Design-Build (DB)"),
    ("Total Building GSF (iNFADS)", 25390),
    ("PRV (iNFADS)", 39507617),
    ("LSH (2.5% of PRV)", 987690),
    ("Locked Total Project Cost ($)", 3527753),
    ("Locked TPC ($000, printed)", 3528),
    ("Prepared By", "Anthony L. Potter, Facilities Planner, MCIPAC G-F PPE"),
    ("Signing Officer (PWO)", "CDR Justus K. O'Connor"),
    ("Government Client", "Bil Hawkins, P.E., PMP, Deputy PWO"),
    ("Reviewer", "Robert W. Kaye, RA, Planning Director NAVFAC FEAD"),
]
for i, (k, v) in enumerate(rows, start=3):
    c.cell(i, 1, k).font = B
    c.cell(i, 2, v)
c.column_dimensions["A"].width = 38
c.column_dimensions["B"].width = 80

# BACKUP
bk = wb.create_sheet("BACKUP")
bk["A1"] = "BACKUP - Locked TPC back-solver (drives BLOCK9 face)"; bk["A1"].font = Font(bold=True, size=12)
bk["A3"] = "INPUTS (locked)"; bk["A3"].font = B; bk["A3"].fill = HDR
inputs = [
    ("Locked TPC ($)", 3527753),
    ("GSF (iNFADS, locked Quantity)", 25390),
    ("PRV ($, iNFADS)", 39507617),
    ("LSH ($, 2.5% of PRV)", "=B6*0.025"),
    ("Multiplier (1.10 x 1.08 x 1.04)", 1.23552),
]
for i, (k, v) in enumerate(inputs, start=4):
    bk.cell(i, 1, k).font = B
    cell = bk.cell(i, 2, v); cell.fill = LOCK
bk["B4"].number_format = '"$"#,##0'
bk["B6"].number_format = '"$"#,##0'
bk["B7"].number_format = '"$"#,##0'
bk["B5"].number_format = '#,##0'
bk["B8"].number_format = '0.00000'

bk["A10"] = "PAX ENTRY (what Anthony types into Block 9)"; bk["A10"].font = B; bk["A10"].fill = HDR
bk["A11"] = "Unit Cost ($/SF)"; bk["A11"].font = B
bk["B11"] = "=ROUND(B4/B5/B8,2)"; bk["B11"].number_format = '"$"#,##0.00'
bk["A12"] = "Items Table Subtotal (raw $) = GSF x UC"; bk["A12"].font = B
bk["B12"] = "=B5*B11"; bk["B12"].number_format = '"$"#,##0'
bk["A13"] = "Items Table Subtotal ($000)"; bk["A13"].font = B
bk["B13"] = "=ROUND(B12/1000,0)"; bk["B13"].number_format = '#,##0'

bk["A15"] = "PAX ROLLUP (computed from items table subtotal)"; bk["A15"].font = B; bk["A15"].fill = HDR
rollup = [
    ("Subtotal ($)",                "=B12"),
    ("Contingency (10.0%)",         "=B16*0.10"),
    ("Total Contract Cost",         "=B16+B17"),
    ("SIOH (8.0%)",                 "=B18*0.08"),
    ("Total Funded Cost",           "=B18+B19"),
    ("Design-Build Design (4.0%)",  "=B20*0.04"),
    ("Total Project Cost ($)",      "=B20+B21"),
    ("Total Project Cost ($000)",   "=ROUND(B22/1000,0)"),
    ("Planning and Design (6.0%) (NON ADD)", "=B22*0.06"),
]
for i, (k, v) in enumerate(rollup, start=16):
    bk.cell(i, 1, k).font = B
    cc = bk.cell(i, 2, v); cc.number_format = '"$"#,##0'
bk["B23"].number_format = '#,##0'

bk["A26"] = "RECONCILIATION"; bk["A26"].font = B; bk["A26"].fill = HDR
bk["A27"] = "Workbook TPC ($)"
bk["B27"] = "=B22"; bk["B27"].number_format = '"$"#,##0'
bk["A28"] = "Locked TPC ($)"
bk["B28"] = "=B4"; bk["B28"].number_format = '"$"#,##0'
bk["A29"] = "Delta ($)"
bk["B29"] = "=B27-B28"; bk["B29"].number_format = '"$"#,##0;[Red]"$"-#,##0'
bk["A30"] = "Workbook TPC ($000)"
bk["B30"] = "=B23"; bk["B30"].number_format = '#,##0'
bk["A31"] = "Locked TPC ($000, printed)"
bk["B31"] = "=ROUND(B4/1000,0)"; bk["B31"].number_format = '#,##0'
bk["A32"] = "Delta ($000) - MUST BE 0"
bk["B32"] = "=B30-B31"; bk["B32"].number_format = '#,##0;[Red]-#,##0'; bk["B32"].font = B

bk.column_dimensions["A"].width = 44
bk.column_dimensions["B"].width = 18

# BLOCK9
b9 = wb.create_sheet("BLOCK9")
b9["A1"] = "9. Cost Estimates"; b9["A1"].font = Font(bold=True, size=12)
hdr = ["Item", "UM", "Quantity", "Unit Cost", "Cost ($000)"]
for j, h in enumerate(hdr, start=1):
    cc = b9.cell(3, j, h); cc.font = B; cc.fill = HDR; cc.border = BOX; cc.alignment = Alignment(horizontal="center")

b9["A4"] = "PRIMARY FACILITY"; b9["A4"].font = B
b9["E4"] = "=BACKUP!B13"; b9["E4"].number_format = '#,##0'; b9["E4"].font = B

b9["A5"] = "  AUTO ORGANIZATIONAL SHOP CAB FAC#3270 (Conversion / Alteration)"
b9["B5"] = "SF"; b9["B5"].alignment = Alignment(horizontal="center")
b9["C5"] = "=BACKUP!B5"; b9["C5"].number_format = '#,##0'; b9["C5"].alignment = Alignment(horizontal="right")
b9["D5"] = "=BACKUP!B11"; b9["D5"].number_format = '#,##0.00'; b9["D5"].alignment = Alignment(horizontal="right")
b9["E5"] = "=ROUND(C5*D5/1000,0)"; b9["E5"].number_format = '#,##0'; b9["E5"].alignment = Alignment(horizontal="right")

# Apply borders to data rows
for r in range(3, 6):
    for ccol in range(1, 6):
        b9.cell(r, ccol).border = BOX

b9["A7"] = "Subtotal"; b9["A7"].font = B
b9["E7"] = "=BACKUP!B13"; b9["E7"].number_format = '#,##0'; b9["E7"].font = B
b9["A8"] = "Contingency (10.0%)"
b9["E8"] = "=ROUND(BACKUP!B17/1000,0)"; b9["E8"].number_format = '#,##0'
b9["A9"] = "Total Contract Cost"; b9["A9"].font = B
b9["E9"] = "=ROUND(BACKUP!B18/1000,0)"; b9["E9"].number_format = '#,##0'; b9["E9"].font = B
b9["A10"] = "Supervision, Inspection and Overhead (8.0%)"
b9["E10"] = "=ROUND(BACKUP!B19/1000,0)"; b9["E10"].number_format = '#,##0'
b9["A11"] = "Total Funded Cost"; b9["A11"].font = B
b9["E11"] = "=ROUND(BACKUP!B20/1000,0)"; b9["E11"].number_format = '#,##0'; b9["E11"].font = B
b9["A12"] = "Design-Build - Design Cost (4.0%)"
b9["E12"] = "=ROUND(BACKUP!B21/1000,0)"; b9["E12"].number_format = '#,##0'
b9["A13"] = "TOTAL PROJECT COST"; b9["A13"].font = B; b9["A13"].fill = HDR
b9["E13"] = "=BACKUP!B23"; b9["E13"].number_format = '#,##0'; b9["E13"].font = B; b9["E13"].fill = HDR
b9["A14"] = "Planning and Design (6.0%) (NON ADD)"
b9["E14"] = "=ROUND(BACKUP!B24/1000,0)"; b9["E14"].number_format = '"("#,##0")"'
b9["A15"] = "Equipment from Other Appropriations (NON ADD)"
b9["E15"] = 0; b9["E15"].number_format = '"("#,##0")"'

b9["A17"] = "Classification of Work"; b9["A17"].font = B
b9["A18"] = "  Construction"
b9["E18"] = "=E13"; b9["E18"].number_format = '#,##0'
b9["A19"] = "Special Interest Codes"; b9["A19"].font = B
b9["A20"] = "  Restoration and Modernization"
b9["E20"] = "=E13"; b9["E20"].number_format = '#,##0'

b9.column_dimensions["A"].width = 60
b9.column_dimensions["B"].width = 6
b9.column_dimensions["C"].width = 12
b9.column_dimensions["D"].width = 12
b9.column_dimensions["E"].width = 14

# BLOCK10
b10 = wb.create_sheet("BLOCK10")
b10["A1"] = "10. Description of Proposed Construction"; b10["A1"].font = Font(bold=True, size=12)
with open("working/3270.txt") as f:
    txt = f.read()
import re
m = re.search(r'10\.\s*Description of Proposed Construction:(.*?)(?=\n11\.\s*Requirement)', txt, re.DOTALL)
narr = m.group(1) if m else ""
narr = re.sub(r'DD1391C? Form.*?\d{2}-\w{3}-\d{2}\s*', '', narr, flags=re.DOTALL)
narr = re.sub(r'1\. Component.*?BU26PPE\d+M\s+[\d,]+\s*\n', '', narr, flags=re.DOTALL)
narr = re.sub(r'10\. Description of Proposed Construction:\(Continued\)\s*\n', '', narr)
narr = narr.strip()
b10["A3"] = narr
b10["A3"].alignment = Alignment(wrap_text=True, vertical="top")
b10.column_dimensions["A"].width = 110
b10.row_dimensions[3].height = 800

# RATES
rt = wb.create_sheet("RATES")
rt["A1"] = "RATES (PAX Associated Costs and project parameters)"; rt["A1"].font = Font(bold=True, size=12)
rt["A3"] = "PAX Associated Costs (pre-loaded by operator)"; rt["A3"].font = B; rt["A3"].fill = HDR
rates = [
    ("Contingency", 0.10, "FSRM program-directed; CRB Guidelines May 2024 (REV #15)"),
    ("Supervision Inspection and Overhead (SIOH)", 0.08, "OCONUS FSRM customer-directed for MCIPAC G-F"),
    ("Design Build (DBD) - REQUIRED", 0.04, "UFC 3-730-01 (2024) - applied after Total Funded Cost"),
    ("Planning and Design (P&D) (NON ADD)", 0.06, "Does not roll into TPC"),
]
for i, (k, v, src) in enumerate(rates, start=4):
    rt.cell(i, 1, k).font = B
    cc = rt.cell(i, 2, v); cc.number_format = "0.0%"
    rt.cell(i, 3, src)
rt["A9"] = "Project parameters"; rt["A9"].font = B; rt["A9"].fill = HDR
params = [
    ("Area Cost Factor (ACF)", 1.85, "UFC 3-701-01 w/Change 7 (25 Jul 2025), Table 4-1, M67400-0004 OCONUS Okinawa"),
    ("LSH Rate (% of PRV)", 0.025, "Local direction at startup"),
    ("JPY/USD Planning Rate", 150.4415, "FY26 budget planning rate"),
    ("JPY/USD H.10 Rate (live)", "", "Verify Federal Reserve H.10 on date of PAX submission"),
    ("Escalation (annual)", 0.021, "OSD FY25 published rate"),
    ("General Requirements", 0.10, "FSRM program-directed"),
]
for i, (k, v, src) in enumerate(params, start=10):
    rt.cell(i, 1, k).font = B
    cc = rt.cell(i, 2, v)
    if isinstance(v, float) and v < 1:
        cc.number_format = "0.0%"
    rt.cell(i, 3, src)
rt.column_dimensions["A"].width = 38
rt.column_dimensions["B"].width = 14
rt.column_dimensions["C"].width = 80

# REFERENCES
rf = wb.create_sheet("REFERENCES")
rf["A1"] = "REFERENCES (authority cites)"; rf["A1"].font = Font(bold=True, size=12)
rf["A3"] = "Document"; rf["A3"].font = B; rf["A3"].fill = HDR
rf["B3"] = "Date / Edition"; rf["B3"].font = B; rf["B3"].fill = HDR
rf["C3"] = "Authority for"; rf["C3"].font = B; rf["C3"].fill = HDR
refs = [
    ("MCO 11000.5", "03 Jun 2016", "Real Property FSRM Program"),
    ("MCO 11000.12", "08 Sep 2014", "Real Property Facilities Manual"),
    ("CRB Guidelines (REV #15)", "May 2024", "Block 9 line-item sequencing"),
    ("UFC 3-701-01 with Change 7", "25 Jul 2025", "PUC and ACF tables, Equation 3-1 PRV formula"),
    ("UFC 3-730-01", "2024", "Design-Build contracting; DB Design 4% applied after TFC"),
    ("JED Cost Estimating Guide", "Nov 2025", "Cost estimating methodology"),
    ("Japan District Design Guide v9", "April 2025", "Okinawa-specific design criteria"),
]
for i, (k, d, src) in enumerate(refs, start=4):
    rf.cell(i, 1, k).font = B
    rf.cell(i, 2, d)
    rf.cell(i, 3, src)
rf["A12"] = "NOT CITED"
rf["A12"].font = B; rf["A12"].fill = HDR
rf["A13"] = "NAVFAC 11010.44E"
rf["B13"] = "INACTIVE"
rf["C13"] = "Inactivated by 1987 issuance; not citable"
rf.column_dimensions["A"].width = 32
rf.column_dimensions["B"].width = 18
rf.column_dimensions["C"].width = 70

# Tab order
order = ["COVER", "BLOCK9", "BLOCK10", "BACKUP", "RATES", "REFERENCES"]
wb._sheets = [wb[n] for n in order]

out = "3270_G_CEPBKUP_BU26PPE74M_POM26_20260319.xlsx"
wb.save(out)
print(f"saved {out}")
